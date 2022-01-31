import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from typing import Dict, Optional
from bannelord_czech.bannelordModules import BannelordModules
import ntpath
import os
import pathlib
import argparse


class ExcelToXml(object):
    prefixes = {
        'Native': 'nat_',
        'SandBox': 'sab_',
        'SandBoxCore': 'sbc_',
        'StoryMode': 'sto_'
    }

    def __init__(self, test: bool = False):
        self.modules: Dict[str, Module] = {}
        self._version = 'e1.5.7'
        self.wb = load_workbook('../installer/translate/translate.xlsx')
        self._languagePath = 'ModuleData/Languages/CZ/'
        self._ignore_sheets = {'Slovník', 'List1'}
        self._test = test

        self._folder = 'E:/Bannelord - čeština/'
        self._foldModules = BannelordModules(folder=f'{self._folder}{self._version}/Modules')

    def transform(self):
        ET.register_namespace('xsi', 'http://www.w3.org/2001/XMLSchema-instance')
        ET.register_namespace('xsd', 'http://www.w3.org/2001/XMLSchema')

        for sheet in self.wb.worksheets:
            title = sheet.title

            if title in self._ignore_sheets:
                continue

            self.modules[title] = Module(sheet)
            prefix = ''

            if self._test:
                prefix = self.prefixes[title]

            for row in range(2, sheet.max_row):
                file = ''
                try:
                    file = f'{self._folder}{self._version}/Modules/{title}/{self._languagePath}' \
                           f'{sheet.cell(row, 2).value.replace(" ", "")}'
                except AttributeError:
                    print(sheet.cell(row, 2))
                string_id = sheet.cell(row, 3).value
                translate = f'{prefix}{sheet.cell(row, 5).value}'
                self.modules[title].change_string(file, string_id, translate)

        self.save()

    def save(self):
        for module in self.modules:
            path = f'translate/Modules/{module}/{self._languagePath}'
            self.modules[module].save_xmls(path)


class Module(object):
    def __init__(self, name):
        self.xmls: Dict[str, "BannelordXml"] = {}
        self.name = name

    def change_string(self, file: str, string_id: str, translate: str):
        if file not in self.xmls:
            if not self.open_xml(file):
                return

        strings = self.xmls[file].get_strings
        text = strings.find(f"string[@id='{string_id}']")

        if text is not None and translate is not None and translate != '':
            text.set('text', str(translate))

    def create_xml(self, file: str):
        self.xmls[file] = BannelordXml()

    def open_xml(self, file) -> bool:
        try:
            tree = ET.parse(file)
            self.xmls[file] = BannelordXml(tree)
            return True
        except FileNotFoundError:
            return False

    def save_xmls(self, path: str):
        for xml in self.xmls:
            file = ntpath.basename(xml)
            self.xmls[xml].save(path, file)


class BannelordXml(object):
    def __init__(self, tree: Optional[ET.ElementTree] = None):
        if tree is None:
            self._base = ET.Element('base')
            self._tags = ET.SubElement(self._base, 'tags')
            self._strings = ET.SubElement(self._base, 'strings')
        else:
            self._tree = tree
            self._base: ET.Element = self._tree.getroot()
            self._tags = self._base.find('tags')

            if self._base.find('strings') is not None:
                self._strings = self._base.find('strings')

    def save(self, path: str, file: str):
        self._base.set('xmlns:xsi', 'http://www.w3.org/2001/XMLSchema-instance')
        self._base.set('xmlns:xsd', 'http://www.w3.org/2001/XMLSchema')

        if not os.path.isdir(path):
            pathlib.Path(path).mkdir(parents=True, exist_ok=True)

        self._tree.write(f'{path}{file}', encoding='utf-8', xml_declaration=True, method='xml')

    @property
    def get_strings(self) -> ET.SubElement:
        return self._strings


class ConvertArguments(object):
    def __init__(self):
        self._test = False

    def parse(self):
        parser = argparse.ArgumentParser()
        parser.add_argument('--test', action='store_true')
        args = parser.parse_args()
        self._test = args.test

    @property
    def test(self):
        return self._test


if __name__ == '__main__':
    convertArguments = ConvertArguments()
    convertArguments.parse()
    excelToXml = ExcelToXml(convertArguments.test)
    excelToXml.transform()
